import openpyxl


def readExcel(filename):
    workbook = openpyxl.load_workbook(filename)
    #sheets = workbook.sheetnames
    worksheet = workbook.get_sheet_by_name(workbook.sheetnames[0])
    groupNumberList = []
    stuIdList = []
    nameList = []
    for cell in list(worksheet.columns)[0]:
        groupNumberList.append(cell.value)
    for cell in list(worksheet.columns)[1]:
        stuIdList.append(str(cell.value))
    for cell in list(worksheet.columns)[2]:
        nameList.append(str(cell.value))
    del groupNumberList[0]
    del stuIdList[0]
    del nameList[0]

    # ####for debug
    # print(groupNumberList)
    # print("\n")
    # print(stuIdList)
    # print("\n")
    # print(nameList)
    # ###
    return groupNumberList, stuIdList, nameList



# # debug
# filename1 = "C:/Users/Yihuan/Desktop/名单.xlsx"
# print(filename1)
# readExcel(filename1)